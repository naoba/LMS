from django.shortcuts import render,get_object_or_404, redirect
from django.http import HttpResponse
from .models import Customer
from django.core.paginator import Paginator
from .forms import CustomerForm, JEForm
from openpyxl import load_workbook, Workbook
import pandas as pd



# Create your views here.

def CustomerLists(request):
    customers_list = Customer.objects.all()
    paginator = Paginator(customers_list, 25)

    page_number = request.GET.get("page")
    customers = paginator.get_page(page_number)

    context = {'customers':customers}  
    return render(request, 'customers/customerlist.html', context)

def CustomerView(request, pk):

    customer = Customer.objects.get(id=pk)

    context = {'customer':customer}  
    return render(request, 'customers/customerview.html', context)

def CustomerCreate(request):

    if request.method == "POST":
        form = CustomerForm(request.POST,request.FILES)
        if form.is_valid():
            # handle_uploaded_file(request.FILES["file"])
            item = form.save(commit=False)
            item.save()
            return redirect('customerview', pk=item.pk)
    else:
        form = CustomerForm()
    return render(request, 'customers/customercreate.html', {'form': form, })


def CustomerEdit(request, pk):

    customer = get_object_or_404(Customer, id=pk)
    if request.method == "POST":
        form = CustomerForm(request.POST, instance=customer,)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.save()
            return redirect('customerview', pk=customer.pk)
    else:
        form = CustomerForm(instance=customer)

    return render(request, 'customers/customercreate.html', {'form': form,'title':pk})

def JsontoExcel(request):
    if request.method == "POST":
        form = JEForm(request.POST,request.FILES)
        if form.is_valid():
            JE = form.save(commit=False)
            JE.save()
            # print(JE.FullPath())
            import json
            with open(JE.FullPath()) as f:
                alldata = json.load(f)
                data= alldata["data"]
                wb = Workbook()
                sheet = wb.active
                ref=[]
                for i in range(len(data)):
                    n=str(i+2)
                    sheet["A"+n]= data[i]["referral_name"]
                    sheet["B"+n]= data[i]["dept_name"] 
                    ref.append(data[i]["referral_name"])

                unique_list = pd.Series(ref).drop_duplicates().tolist()
                summary = wb.create_sheet("Summary", 1)
                for i in range(len(unique_list)):
                     n=str(i+2)
                     summary["A"+n] = unique_list[i]
                     summary["B"+n] = ref.count(unique_list[i])

                wb.save(filename="hello_world.xlsx")              
                
       
            return render(request, 'customers/jsontoexcel.html', {"excel_file": JE.JEupload,"form": form })


    else:
        form = JEForm()
        return render(request, 'customers/jsontoexcel.html', {"form": form})

